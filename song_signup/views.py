import logging
import traceback
from functools import wraps
import random
import time
import csv
import io
from contextlib import redirect_stdout

import constance
from constance import config
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.core.management import call_command
from django.db import transaction
from django.db.utils import IntegrityError
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from flags.state import enable_flag, disable_flag, flag_disabled, flag_enabled
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from titlecase import titlecase
from django.core.exceptions import ValidationError
from django.db.models import Count
from .forms import TickchakUploadForm
from .tasks import get_lyrics
from .models import (
    GroupSongRequest,
    SongLyrics,
    SongRequest,
    Singer,
    SongSuggestion,
    TicketOrder,
    SING_SKU,
    ATTN_SKU,
    TicketsDepleted,
    AlreadyLoggedIn,
    CurrentGroupSong,
    TriviaQuestion,
    TriviaResponse,
    Celebration
)
from .serializers import (
    SongSuggestionSerializer,
    SongRequestSerializer,
    SingerSerializer,
    SongRequestLineupSerializer,
    GroupSongRequestLineupSerializer,
    LyricsSerializer,
    TriviaQuestionSerializer,
    TriviaResponseSerializer,
    RaffleWinnerSerializer
)
from twist.utils import format_commas

logger = logging.getLogger(__name__)

class TwistApiException(Exception):
    pass


def bwt_login_required(login_url, singer_only=False):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated or (singer_only and request.user.is_audience
                                                     and not request.user.raffle_winner):
                return redirect(login_url)

            return view_func(request, *args, **kwargs)

        return wrapper

    return decorator


def superuser_required(login_url):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_superuser:
                return redirect(login_url)

            return view_func(request, *args, **kwargs)

        return wrapper

    return decorator


def name_to_username(first_name, last_name):
    return f'{first_name.lower()}_{last_name.lower()}'


@bwt_login_required('login')
def home(request):
    song = request.GET.get('song')
    is_group_song = request.GET.get('is-group-song') == 'true'
    return render(request, 'song_signup/home.html', {
        "new_song": song,
        "is_group_song": is_group_song
    })


def spotlight_data(request):
    current_song, is_group_song = _get_current_song()

    if is_group_song:
        next_song = SongRequest.objects.current_song()
    else:
        next_song = SongRequest.objects.next_song()

    return JsonResponse(
        {
            "current_song": current_song and current_song.basic_data,
            "next_song": next_song and next_song.basic_data,
        })


def next_singer(request):
    """
    Return username of next singer
    """
    next_song = SongRequest.objects.current_song()
    username = next_song.singer.username if next_song is not None else None

    return JsonResponse({
        "next_singer": username
    })


def dashboard_data(request):
    singer = request.user
    user_next_song = singer.next_song

    return JsonResponse({"user_next_song": user_next_song and user_next_song.basic_data,
                         "raffle_winner_already_sang": singer.raffle_winner_already_sang})


def _sanitize_string(name, title=False):
    sanitized = ' '.join(word.capitalize() for word in name.split())
    return titlecase(sanitized) if title else sanitized


@bwt_login_required('login', singer_only=True)
def add_song_request(request):
    current_user = request.user

    if request.method == 'POST':
        song_name = _sanitize_string(request.POST['song-name'], title=True)
        musical = _sanitize_string(request.POST['musical'], title=True)
        notes = request.POST.get('notes')
        partners = request.POST.getlist('partners')
        approve_duplicate = request.POST.get('approve-duplicate')
        crowd_pleaser = bool(request.POST.get('suggested_by'))

        song_request = SongRequest.objects.filter(song_name=song_name, musical=musical).first()
        if not song_request or approve_duplicate:
            with transaction.atomic(): # Abort object creation if validation fails
                # Raffle winner add their songs as standby - to be spotlit when Shani decides.
                new_song_request = SongRequest.objects.create(song_name=song_name, musical=musical,
                                                              singer=current_user, notes=notes,
                                                              standby=current_user.raffle_winner,
                                                              crowd_pleaser=crowd_pleaser)
                try:
                    new_song_request.partners.set(partners) # Runs validation function using signal
                except ValidationError as e:
                    return JsonResponse({"error": e.message}, status=400)

            Singer.ordering.calculate_positions()
            SongSuggestion.objects.check_used_suggestions()

            return JsonResponse({
                'requested_song': new_song_request.song_name,
            })

        else:
            if current_user in song_request.partners.all():
                return JsonResponse({"error": f"Apparently, {song_request.singer} already signed you up for this song"},
                                    status=400)
            elif song_request.singer == current_user:
                return JsonResponse({"error": "You already signed up with this song tonight"}, status=400)
            else:
                return JsonResponse({"duplicate": True}, status=202)


@api_view(["GET"])
def get_current_songs(request):
    singer = request.user
    serialized = SongRequestSerializer(singer.pending_songs, many=True, read_only=True)
    return Response(serialized.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def get_suggested_songs(request):
    suggestions = SongSuggestion.objects.annotate(
        num_votes=Count('voters')
    ).order_by('is_used', '-num_votes', '-request_time')
    
    serialized = SongSuggestionSerializer(suggestions, many=True, read_only=True, context={'request': request})
    return Response(serialized.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def get_current_user(request):
    serialized = SingerSerializer(request.user, read_only=True)
    return Response(serialized.data, status=status.HTTP_200_OK)


@api_view(["POST"])
def toggle_vote(request, suggestion_id):
    """
    Toggle a user's vote for a song suggestion.
    """
    try:
        suggestion = SongSuggestion.objects.get(pk=suggestion_id)
        user = request.user

        if not user.is_authenticated:
            return Response({'error': 'User must be authenticated to vote'}, 
                          status=status.HTTP_401_UNAUTHORIZED)

        if suggestion.user_voted(user):
            suggestion.voters.remove(user)
            voted = False
        else:
            suggestion.voters.add(user)
            voted = True

        return Response({
            'voted': voted
        }, status=status.HTTP_200_OK)

    except SongSuggestion.DoesNotExist:
        return Response({'error': f'Song suggestion with ID {suggestion_id} does not exist'}, 
                       status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
def get_drinking_words(request):
    drinking_words = constance.config.DRINKING_WORDS
    return Response({'drinking_words': drinking_words.split(';') if drinking_words else []},
                    status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAdminUser])
def get_passcode(request):
    passcode = constance.config.PASSCODE
    return Response({'passcode': passcode}, status=status.HTTP_200_OK)


@api_view(["GET"])
def get_song(request, song_pk):
    try:
        song_request = SongRequest.objects.get(pk=song_pk)
        serialized = SongRequestSerializer(song_request, read_only=True)
        return Response(serialized.data, status=status.HTTP_200_OK)
    except SongRequest.DoesNotExist:
        return Response({'error': f"Song with ID {song_pk} does not exist"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def get_lineup(request):
    song_requests = SongRequest.objects.filter(position__isnull=False, skipped=False).order_by('position')
    current_song, is_group_song = _get_current_song()

    if is_group_song:
        current_song_data = GroupSongRequestLineupSerializer(current_song).data
        next_songs_data = SongRequestLineupSerializer(song_requests, many=True).data
    else:
        current_song_data = SongRequestLineupSerializer(current_song).data
        next_songs_data = SongRequestLineupSerializer(song_requests[1:], many=True).data

    return Response({
        'current_song': current_song_data,
        'next_songs': next_songs_data
    }, status=status.HTTP_200_OK)


@api_view(["PUT"])
def update_song(request):
    song_id = request.data['song_id']
    song_name = _sanitize_string(request.data['song_name'], title=True)
    musical = _sanitize_string(request.data['musical'], title=True)
    notes = request.data.get('notes')
    partner_ids = request.data.get('partners', [])
    partners = list(Singer.objects.filter(id__in=partner_ids))

    if not song_name:
        return Response({'error': f"Song name can not empty"}, status=status.HTTP_400_BAD_REQUEST)

    if not musical:
        return Response({'error': f"Musical name can not empty"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        song_request = SongRequest.objects.get(pk=song_id)
        if song_request.song_name != song_name:
            song_request.found_music = False
            song_request.default_lyrics = False
            song_request.to_alon = None
        song_request.song_name = song_name
        song_request.musical = musical
        song_request.notes = notes
        song_request.partners.set(partners)
        song_request.save()

        serialized = SongRequestSerializer(song_request, read_only=True)
        return Response(serialized.data, status=status.HTTP_200_OK)

    except ValidationError as e:
        return JsonResponse({"error": e.message}, status=status.HTTP_400_BAD_REQUEST)

    except SongRequest.DoesNotExist:
        return Response({'error': f"Song with ID {song_id} does not exist"}, status=status.HTTP_400_BAD_REQUEST)


@bwt_login_required('login', singer_only=True)
def manage_songs(request):
    return render(request, 'song_signup/manage_songs.html')


@bwt_login_required('login')
def view_suggestions(request):
    return render(request, 'song_signup/view_suggestions.html')


@bwt_login_required('login', singer_only=True)
def add_song(request):
    return render(request, 'song_signup/add_song.html', {'possible_partners': _get_possible_partners(request)})

def _get_possible_partners(request):
    superusers = Singer.objects.filter(is_superuser=True).all()
    partner_options = Singer.objects.filter(is_active=True).exclude(pk=request.user.pk).exclude(is_superuser=True).order_by(
        'first_name')
    return list(superusers) + list(partner_options)

@api_view(["GET"])
def get_possible_partners(request):
    partners = _get_possible_partners(request)
    serialized = SingerSerializer(partners, many=True)
    return Response(serialized.data, status=status.HTTP_200_OK)


def _sort_lyrics(song: SongRequest | GroupSongRequest):
    """
    Sort the lyrics based on our best guess of how well they match
    Tried a few algorithms here but this relatively simple one worked best:

    1) If default is selected order first
    2) Songs with exact name matches
    3) Songs with left matches (input song name is included in full song name)
    4) Songs with right matches (full song name is included in input song name)
    5) Everything else

    Within each group just order by ID for consistency.
    """
    if not song:
        return

    lyrics = song.lyrics.order_by('id').all()

    default = [lyric for lyric in lyrics if lyric.default]
    lyrics = [lyric for lyric in lyrics if lyric not in default]

    exact_matches = [lyric for lyric in lyrics if song.song_name.lower() == lyric.song_name.lower()]
    lyrics = [lyric for lyric in lyrics if lyric not in exact_matches]

    left_matches = [lyric for lyric in lyrics if song.song_name.lower() in lyric.song_name.lower()]
    lyrics = [lyric for lyric in lyrics if lyric not in left_matches]

    right_matches = [lyric for lyric in lyrics if lyric.song_name.lower() in song.song_name.lower()]
    lyrics = [lyric for lyric in lyrics if lyric not in right_matches]

    return default + exact_matches + left_matches + right_matches + lyrics


@superuser_required('login')
def lyrics(request, song_pk):
    try:
        song_request = SongRequest.objects.get(pk=song_pk)
    except SongRequest.DoesNotExist:
        return JsonResponse({'error': f"Song with ID {song_pk} does not exist"}, status=status.HTTP_400_BAD_REQUEST)

    lyrics = _sort_lyrics(song_request)
    return render(request, 'song_signup/lyrics.html', {"lyrics": lyrics and lyrics[0], "song": song_request})


@superuser_required('login')
def group_lyrics(request, song_pk):
    try:
        song_request = GroupSongRequest.objects.get(pk=song_pk)
    except GroupSongRequest.DoesNotExist:
        return JsonResponse({'error': f"Group song with ID {song_pk} does not exist"},
                            status=status.HTTP_400_BAD_REQUEST)

    lyrics = _sort_lyrics(song_request)
    return render(request, 'song_signup/lyrics.html', {"lyrics": lyrics and lyrics[0], "group_song": song_request})


@superuser_required('login')
def force_reset_lyrics(request, song_pk):
    get_lyrics.delay(song_id=song_pk)
    time.sleep(1)
    return redirect('alternative_lyrics', song_pk)

@superuser_required('login')
def force_reset_lyrics_group(request, song_pk):
    get_lyrics.delay(group_song_id=song_pk)
    time.sleep(1)
    return redirect('alternative_group_lyrics', song_pk)


@superuser_required('login')
def lyrics_by_id(request, lyrics_id):
    try:
        lyrics = SongLyrics.objects.get(id=lyrics_id)
    except SongLyrics.DoesNotExist:
        return JsonResponse({'error': f"Lyrics with ID {lyrics_id} does not exist"}, status=status.HTTP_400_BAD_REQUEST)

    return render(request, 'song_signup/lyrics.html', {
        "lyrics": lyrics,
        "song": lyrics.song_request,
        "group_song": lyrics.group_song_request
    })


@superuser_required('login')
def alternative_lyrics(request, song_pk):
    try:
        song_request = SongRequest.objects.get(pk=song_pk)
    except GroupSongRequest.DoesNotExist:
        return JsonResponse({'error': f"Group song with ID {song_pk} does not exist"},
                            status=status.HTTP_400_BAD_REQUEST)

    lyrics = _sort_lyrics(song_request)
    return render(request, 'song_signup/alternative_lyrics.html', {"lyrics": lyrics, "song": song_request})


@superuser_required('login')
def alternative_group_lyrics(request, song_pk):
    try:
        song_request = GroupSongRequest.objects.get(pk=song_pk)
    except GroupSongRequest.DoesNotExist:
        return JsonResponse({'error': f"Group song with ID {song_pk} does not exist"},
                            status=status.HTTP_400_BAD_REQUEST)

    lyrics = _sort_lyrics(song_request)
    return render(request, 'song_signup/alternative_lyrics.html', {"lyrics": lyrics, "song": song_request})


@api_view(["PUT"])
def default_lyrics(request):
    lyrics_id = request.data['lyricsId']

    lyric = SongLyrics.objects.get(id=lyrics_id)

    if not (request.user.is_superuser or (lyric.song_request and lyric.song_request.singer == request.user)):
        return Response({"error": "Only original singer can set default lyrics"}, status=status.HTTP_401_UNAUTHORIZED)

    lyric.default = True
    lyric.save()

    if lyric.song_request:
        lyric.song_request.default_lyrics = True
        lyric.song_request.save()

    if lyric.group_song_request:
        lyric.group_song_request.default_lyrics = True
        lyric.group_song_request.save(get_lyrics=False)

    return Response({'is_group_song': bool(lyric.group_song_request)}, status=status.HTTP_200_OK)


def _get_current_song():
    """
    Return current_song, is_group_song
    """
    curr_group_song = CurrentGroupSong.objects.all().first()
    if curr_group_song and curr_group_song.is_active:
        return curr_group_song.group_song, True
    else:
        return SongRequest.objects.get_spotlight() or SongRequest.objects.current_song(), False


@api_view(["GET"])
def get_current_lyrics(request):
    current, is_group_song = _get_current_song()
    lyrics = _sort_lyrics(current)
    serialized = LyricsSerializer(lyrics[0] if lyrics else None, many=False, read_only=True,
                                  context={'is_group_song': is_group_song})
    return Response(serialized.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def demo_current_lyrics(request):
    group_song = GroupSongRequest.objects.get(
        song_name="Be Our Guest",
        musical="Beauty And The Beast",
    )

    lyrics = group_song.lyrics.get(default=True)
    serialized = LyricsSerializer(lyrics, many=False, read_only=True, context={'is_group_song': False})
    return Response(serialized.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def get_active_question(request):
    active_question = TriviaQuestion.objects.filter(is_active=True).first()
    if active_question:
        serialized = TriviaQuestionSerializer(active_question)
        return Response(serialized.data, status=status.HTTP_200_OK)
    else:
        return Response({}, status=status.HTTP_200_OK)


@api_view(["POST"])
def select_trivia_answer(request):
    answer_id = request.data['answer-id']
    user = request.user
    active_question = TriviaQuestion.objects.filter(is_active=True).first()

    if active_question:
        if not user.trivia_responses.filter(question=active_question):
            TriviaResponse.objects.create(user=user, choice=answer_id, question=active_question)
            return Response({}, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': "User already selected a question"},
                            status=status.HTTP_409_CONFLICT)
    else:
        return Response({'error': "No active question"},
                        status=status.HTTP_400_BAD_REQUEST)

@api_view(["GET"])
def get_selected_answer(request):
    user = request.user
    active_question = TriviaQuestion.objects.filter(is_active=True).first()
    if active_question:
        trivia_response = user.trivia_responses.filter(question=active_question).first()
        if trivia_response:
            serialized = TriviaResponseSerializer(trivia_response)
            return Response(serialized.data, status=status.HTTP_200_OK)

    return Response({'error': "User hasn't selected an answer for the active question"},
                    status=status.HTTP_404_NOT_FOUND)


@superuser_required('login')
def deactivate_trivia(request):
    TriviaQuestion.objects.all().update(is_active=False)
    return redirect('admin/song_signup/triviaquestion')


@superuser_required('login')
def start_group_song(request):
    CurrentGroupSong.start_song()
    return redirect('admin/song_signup/groupsongrequest')


@superuser_required('login')
def end_group_song(request):
    CurrentGroupSong.end_song()
    if "redirect_songrequest" in request.GET:
        return redirect('admin/song_signup/songrequest')

    return redirect('admin/song_signup/groupsongrequest')


@superuser_required('login')
def end_spotlight(request):
    SongRequest.objects.remove_spotlight()
    return redirect('admin/song_signup/songrequest')


@bwt_login_required('login')
def live_lyrics(request):
    return render(request, 'song_signup/live_lyrics.html')


@bwt_login_required('login')
def song_voting(request):
    return render(request, 'song_signup/song_voting.html')


@bwt_login_required('login')
def lineup(request):
    return render(request, 'song_signup/lineup.html')


@superuser_required('login')
def demo_live_lyrics(request):
    return render(request, 'song_signup/live_lyrics.html', {'demo_mode': True})


@superuser_required('login')
def start_raffle(request):
    winner = Singer.objects.filter(is_audience=True, raffle_participant=True, raffle_winner=False, is_active=True).order_by('?').first()
    if winner:
        winner.raffle_winner = True
        winner.active_raffle_winner = True
        winner.save()
    else:
        winner = "NO RAFFLE PARTICIPANTS"

    request.session['raffle_winner'] = str(winner)
    return redirect(f'admin/song_signup/songrequest')

@superuser_required('login')
def end_raffle(request):
    for active_winner in Singer.objects.filter(active_raffle_winner=True):
        active_winner.active_raffle_winner = False
        active_winner.save()

    del request.session['raffle_winner']
    return redirect(f'admin/song_signup/songrequest')

@bwt_login_required('login')
def toggle_raffle_participation(request):
    user = request.user

    user.raffle_participant = not user.raffle_participant
    user.save()
    return redirect('home')

@api_view(["GET"])
def get_active_raffle_winner(request):
    active_winner = Singer.objects.filter(active_raffle_winner=True).first()
    if active_winner:
        serialized = RaffleWinnerSerializer(active_winner)
        return Response(serialized.data, status=status.HTTP_200_OK)
    else:
        return Response({}, status=status.HTTP_200_OK)

@bwt_login_required('login')
def suggest_song(request):
    current_user = request.user

    if request.method == 'POST':
        song_name = _sanitize_string(request.POST['song-name'], title=True)
        musical = _sanitize_string(request.POST['musical'], title=True)
        suggested_by = str(current_user)

        if request.POST.get('group-song') == 'on':
            GroupSongRequest.objects.create(song_name=song_name, musical=musical, suggested_by=suggested_by)
            return redirect(f"{reverse('home')}?song={song_name}&is-group-song=true")
        else:
            _, created = SongSuggestion.objects.get_or_create(song_name=song_name, musical=musical,
                                                              defaults={'suggested_by': current_user})
            if created:
                SongSuggestion.objects.check_used_suggestions()
            return redirect('song_voting')

    else:
        return render(request, 'song_signup/suggest_song.html')


def logout(request):
    user = request.user

    if not user.is_superuser:
        user.is_active = False

    user.save()
    auth_logout(request)
    Singer.ordering.calculate_positions()
    return redirect('login')


def faq(request):
    return render(request, 'song_signup/faq.html')


def tip_us(request):
    return render(request, 'song_signup/tip_us.html')


def _get_order(order_id, ticket_type):
    if config.FREEBIE_TICKET and order_id == config.FREEBIE_TICKET:
        return TicketOrder.objects.get_or_create(order_id=config.FREEBIE_TICKET,
                                                 event_sku=config.EVENT_SKU,
                                                 event_name='FREEBIE-ORDER',

                                                 num_tickets=-1,
                                                 customer_name='FREEBIE_ORDER',
                                                 ticket_type=SING_SKU,
                                                 is_freebie=True)[0]

    try:
        return TicketOrder.objects.get(order_id=order_id,
                                       event_sku=config.EVENT_SKU,
                                       ticket_type=ticket_type)
    except (TicketOrder.DoesNotExist, ValueError):
        ticket_str = 'a singer' if ticket_type == SING_SKU else 'an audience'
        raise TwistApiException(f"We can't find {ticket_str} ticket with that order number. Maybe you have a typo? "
                                "The number appears in the title of the tickets email")


def _login_new_singer(first_name, last_name, no_image_upload, order_id, uploaded_image):
    """
    Singer and audience are different ticket orders entries, both with the same order number
    """
    ticket_order = _get_order(order_id, SING_SKU)
    try:
        singer = Singer.objects.create_user(
            name_to_username(first_name, last_name),
            first_name=first_name,
            last_name=last_name,
            is_staff=False,
            no_image_upload=no_image_upload,
            ticket_order=ticket_order,
            selfie=uploaded_image
        )

    except (TicketsDepleted, AlreadyLoggedIn) as e:
        raise TwistApiException(str(e))
    return singer


def _login_existing_singer(first_name, last_name, no_image_upload, uploaded_image):
    try:
        singer = Singer.objects.get(first_name=first_name, last_name=last_name, is_audience=False)
        singer.is_active = True
        singer.no_image_upload = no_image_upload
        if uploaded_image:
            singer.selfie = uploaded_image
        singer.save()
        Singer.ordering.calculate_positions()
        return singer
    except Singer.DoesNotExist:
        raise TwistApiException("The name that you logged in with previously does not match your current one")


def _login_new_audience(first_name, last_name, no_image_upload, order_id, uploaded_image):
    ticket_order = _get_order(order_id, ATTN_SKU)
    try:
        singer = Singer.objects.create_user(
            name_to_username(first_name, last_name),
            first_name=first_name,
            last_name=last_name,
            is_staff=False,
            is_audience=True,
            ticket_order=ticket_order,
            no_image_upload=no_image_upload,
            selfie=uploaded_image
        )
    except (TicketsDepleted, AlreadyLoggedIn) as e:
        raise TwistApiException(str(e))
    return singer


def  _login_existing_audience(first_name, last_name, no_image_upload, uploaded_image):
    try:
        audience = Singer.objects.get(first_name=first_name, last_name=last_name, is_audience=True)
        audience.is_active = True
        audience.no_image_upload = no_image_upload
        if uploaded_image:
            audience.selfie = uploaded_image
        audience.save()
        return audience
    except Singer.DoesNotExist:
        raise TwistApiException("The name that you logged in with previously does not match your current one")


def login(request):
    # This is the root endpoint. If already logged in, go straight to home.
    constants_chosen = bool(config.PASSCODE) and bool(config.EVENT_SKU)
    if constants_chosen and request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        try:
            ticket_type = request.POST['ticket-type']
            first_name = _sanitize_string(request.POST['first-name'])
            last_name = _sanitize_string(request.POST['last-name'])
            logged_in = request.POST.get('logged-in') == 'on'
            passcode = _sanitize_string(request.POST['passcode'])
            order_id = request.POST['order-id']
            no_image_upload = request.POST.get('no-upload') == 'on'
            photo = request.FILES.get('upload-image')
            selfie = request.FILES.get('upload-selfie')

            uploaded_image = photo or selfie

            if passcode.lower() != config.PASSCODE.lower():
                raise TwistApiException("Wrong passcode - Shani will reveal tonight's passcode at the event")

            if ticket_type == 'audience':
                audience = _login_existing_audience(first_name, last_name, no_image_upload, uploaded_image) if logged_in else (
                    _login_new_audience(first_name, last_name, no_image_upload, order_id, uploaded_image))
                auth_login(request, audience)
                return JsonResponse({'success': True}, status=200)

            elif ticket_type == 'singer':
                singer = _login_existing_singer(first_name, last_name, no_image_upload, uploaded_image) if logged_in else (
                    _login_new_singer(first_name, last_name, no_image_upload, order_id, uploaded_image))
                auth_login(request, singer)
                return JsonResponse({'success': True}, status=200)
            else:
                raise TwistApiException("Invalid ticket type")

        except Exception as e:
            logger.exception("Exception: ")
            if isinstance(e, TwistApiException):
                msg = str(e)
            else:
                msg = "An unexpected error occurred (you can blame Alon..) Refreshing the page might help"
            return JsonResponse({'error': msg}, status=400)

    return render(request, 'song_signup/login.html', context={'evening_started': constants_chosen})


@bwt_login_required('login', singer_only=True)
def delete_song(request, song_pk):
    SongRequest.objects.filter(pk=song_pk).delete()
    Singer.ordering.calculate_positions()
    SongSuggestion.objects.check_used_suggestions()
    return HttpResponse()

def _get_current_filename():
    """
    Returns file name of format 'bwt-5-1-25', according to the currnet event.
    Derives current event from the last ticket order loaded
    """
    event_name = TicketOrder.objects.filter(is_freebie=False).last().event_name
    event_date = event_name.split('-')[-1].strip().replace('.','-')
    return f"bwt-{event_date}"

def _make_setlist():
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['', 'Singers', 'Song', 'Musical'])

    songs = SongRequest.objects.filter(performance_time__isnull=False).order_by('performance_time')
    group_songs = GroupSongRequest.objects.filter(performance_time__isnull=False).order_by('performance_time')

    all_songs = sorted(
        list(songs) + list(group_songs),
        key=lambda x: x.performance_time
    )

    if len(all_songs) >= 2:
        all_songs[-1], all_songs[-2] = all_songs[-2], all_songs[-1]

    for i, song in enumerate(all_songs, start=1):
        if isinstance(song, SongRequest):
            singers = format_commas([song.singer.get_full_name()] +
                                    [singer.get_full_name() for singer in song.partners.all()])
            writer.writerow([i, singers, song.song_name, song.musical])
        elif isinstance(song, GroupSongRequest):
            writer.writerow([i, 'Group Song', song.song_name, song.musical])

    return buffer.getvalue()

@superuser_required('login')
def reset_database(request):
    download_csv = request.GET.get('csv') == 'true'
    filename = _get_current_filename()

    if download_csv:
        setlist = _make_setlist()
        call_command('dbbackup', output_path=f'./db_backups/complete_evenings/{filename}.psql')
    else:
        call_command('dbbackup')

    call_command('reset_db')
    disable_flag('CAN_SIGNUP')
    disable_flag('STARTED')
    config.PASSCODE = ''
    config.DRINKING_WORDS = ''

    if download_csv:
        response = HttpResponse(setlist, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    return redirect('admin/song_signup/songrequest')


@superuser_required('login')
def enable_signup(request):
    enable_flag('CAN_SIGNUP')
    Singer.ordering.calculate_positions()
    return redirect('admin/song_signup/songrequest')


@superuser_required('login')
def disable_signup(request):
    disable_flag('CAN_SIGNUP')
    config.PASSCODE = ''  # So it won't appear in live lyrics at the end of the evening.
    Singer.ordering.calculate_positions()
    return redirect('admin/song_signup/songrequest')


def signup_disabled(request):
    return JsonResponse({"result": flag_disabled('CAN_SIGNUP')})


@superuser_required('login')
def start_evening(request):
    enable_flag('STARTED')
    return redirect('admin/song_signup/songrequest')


@superuser_required('login')
def end_evening(request):
    disable_flag('STARTED')
    return redirect('admin/song_signup/songrequest')


@superuser_required('login')
def start_boho(request):
    enable_flag('BOHO')
    return redirect('admin/song_signup/songrequest')


@superuser_required('login')
def stop_boho(request):
    disable_flag('BOHO')
    return redirect('admin/song_signup/songrequest')


def evening_started(request):
    return JsonResponse({"started": flag_enabled('STARTED')})


def boho_started(request):
    return JsonResponse({"boho": flag_enabled('BOHO')})


@api_view(["PUT"])
@csrf_exempt
def toggle_shani_ping(request):
    if flag_disabled('SHANI_PING'):
        enable_flag('SHANI_PING')
    else:
        disable_flag('SHANI_PING')
    return JsonResponse({"shani_ping": flag_enabled('SHANI_PING')})


@api_view(["GET"])
@csrf_exempt
def shani_pinged(request):
    return JsonResponse({"shani_ping": flag_enabled('SHANI_PING')})


def recalculate_priorities(request):
    Singer.ordering.calculate_positions()
    return redirect('admin/song_signup/songrequest')


@superuser_required('login')
def upload_tickchak_orders(request):
    if request.method == 'POST' and 'file' in request.FILES:
        try:
            spreadsheet = request.FILES['file']
            event_sku = request.POST['event_sku']
            event_date = request.POST['event_date']

            _process_celebrations(spreadsheet, event_date, event_sku)

            generate_cheat_code = request.POST.get("generate_cheat_code") == "on"
            duplicates_upload = request.POST.get("duplicates_upload") == "on"
            processing_data = _process_tickchak_orders(spreadsheet,
                                                       event_sku,
                                                       event_date,
                                                       generate_cheat_code,
                                                       duplicates_upload
                                                       )
        except Exception as e:
            return render(request, 'song_signup/upload_tickchak_orders.html',
                          {'form': TickchakUploadForm(), 'error_message': traceback.format_exc()})

        return HttpResponse(f"""
{"<h3>ONLY CHANGED NEW ENTRIES. DIDN'T CHANGE CONFIGS</h3>" if duplicates_upload else ''}
<h3>{processing_data['num_orders']} Orders Processed Successfully for event {processing_data['event_name']}</h3>
<ul>
<li> EVENT SKU: {processing_data['event_sku']}</li>
{f"<li> CHEAT CODE: {processing_data['cheat_code']}</li>" if processing_data['cheat_code'] else ""}
<br>
<li> Total tickets: {processing_data['total_tickets']}</li>
<li> Num singing tickets: {processing_data['num_singing_tickets']}</li>
<li> Num audience tickets: {processing_data['num_audience_tickets']}</li>
<br>
<li>Num orders (each person counted once): {processing_data['num_orders']}</li>
<li> Num orders (lines in spreadsheet): {processing_data['num_ticket_orders']}</li>
<br>
<li> Num orders that already existed in DB: {processing_data['num_existing_ticket_orders']}</li>
<li> Num new orders added to DB: {processing_data['num_new_ticket_orders']}</li>
<br>
""")

    return render(request, 'song_signup/upload_tickchak_orders.html', {'form': TickchakUploadForm()})


@transaction.atomic
def _process_orders(spreadsheet_file):
    orders = set()  # Order made by a single person
    num_ticket_orders = 0  # Groups of ticket types within the orders
    num_existing_ticket_orders = 0
    num_new_ticket_orders = 0
    num_singing_tickets = 0
    num_audience_tickets = 0

    worksheet = load_workbook(spreadsheet_file).active
    for row in worksheet.iter_rows(min_row=2, values_only=2):
        order_id, event_sku, event_name, num_tickets, ticket_type, first_name, last_name, phone_number = row
        orders.add(order_id)
        num_ticket_orders += 1

        ticket_order, created = TicketOrder.objects.get_or_create(
            order_id=order_id,
            event_sku=event_sku,
            ticket_type=ticket_type,
            event_name=event_name,
            num_tickets=num_tickets,
            customer_name=' '.join([first_name, last_name]),
            phone_number=phone_number
        )
        if created:
            num_new_ticket_orders += 1
        else:
            num_existing_ticket_orders += 1

        if ticket_type == SING_SKU:
            num_singing_tickets += num_tickets
        else:
            num_audience_tickets += num_tickets

        ticket_order.save()

    return dict(num_orders=len(orders),
                num_ticket_orders=num_ticket_orders,
                num_new_ticket_orders=num_new_ticket_orders,
                num_existing_ticket_orders=num_existing_ticket_orders,
                num_singing_tickets=num_singing_tickets,
                num_audience_tickets=num_audience_tickets
                )


ORDER_ID = 'מספר הזמנה'
FIRST_NAME = 'שם פרטי'
LAST_NAME = 'שם משפחה'
PHONE_NUMBER = 'טלפון'
NUM_TICKETS = 'כמות'
TICKET_DESC = 'כותרת'
CELEBRATING = 'חוגגים'


@transaction.atomic
def _process_tickchak_orders(spreadsheet_file, event_sku, event_date, generate_cheat_code=False,
                             duplicates_upload=False):
    SHEET_NAME = 'כרטיסים'
    event_name = f"Open Mic - Babu Bar - {event_date}"

    sheet_fields = (
       ORDER_ID,
       FIRST_NAME,
       LAST_NAME,
       PHONE_NUMBER,
       NUM_TICKETS,
       TICKET_DESC,
    )

    orders = set()  # Order made by a single person
    num_ticket_orders = 0  # Groups of ticket types within the orders
    num_existing_ticket_orders = 0
    num_new_ticket_orders = 0
    num_singing_tickets = 0
    num_audience_tickets = 0

    sample_order = TicketOrder.objects.filter(event_sku=event_sku).first()

    if duplicates_upload:
        if not sample_order:
            raise ValueError(f"SKU {event_sku} doesn't exist")
        if sample_order.event_name != event_name:
                raise ValueError(f"Entered date {event_date}, but existing objects of SKU {event_sku} "
                                 f"have {sample_order.event_name}")
    elif sample_order:
        raise ValueError(f"SKU already exists, but you didn't mark duplicate upload")

    worksheet = load_workbook(spreadsheet_file).get_sheet_by_name(SHEET_NAME)
    column_names = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    missing_columns = [field for field in sheet_fields if field not in column_names]
    if missing_columns:
        raise ValueError(f"The following columns are missing from the spreadsheet: {missing_columns}")

    column_index_map = {name: idx for idx, name in enumerate(column_names)}

    for row in worksheet.iter_rows(min_row=2, values_only=2):
        order_id = row[column_index_map[ORDER_ID]]
        num_tickets = row[column_index_map[NUM_TICKETS]]
        ticket_desc = row[column_index_map[TICKET_DESC]]
        first_name = row[column_index_map[FIRST_NAME]]
        last_name = row[column_index_map[LAST_NAME]]
        phone_number = row[column_index_map[PHONE_NUMBER]]

        orders.add(order_id)
        num_ticket_orders += 1

        ticket_type = SING_SKU if 'זמר' in ticket_desc else ATTN_SKU

        try:
            ticket_order, created = TicketOrder.objects.get_or_create(
                order_id=order_id,
                event_sku=event_sku,
                ticket_type=ticket_type,
                event_name=event_name,
                num_tickets=num_tickets,
                customer_name=' '.join([first_name, last_name]),
                phone_number=phone_number
            )
        except IntegrityError:
            raise ValueError(
                f"A ticket order with order_id={order_id}, event_sku={event_sku}, and ticket_type={ticket_type}"
                f" already exists, but it looks like you're trying to insert an order that has these fields, but other fields "
                f"that are different. Is the event date formatting off? Or did you change some data in the spreadsheet?"
            )
        if created:
            num_new_ticket_orders += 1
        else:
            num_existing_ticket_orders += 1

        if ticket_type == SING_SKU:
            num_singing_tickets += num_tickets
        else:
            num_audience_tickets += num_tickets

        ticket_order.save()

    if not duplicates_upload:
        config.EVENT_SKU = event_sku

        if generate_cheat_code:
            cheat_code = str(random.randint(100000, 999999))
            config.FREEBIE_TICKET = cheat_code
        else:
            cheat_code = None
    else:
        cheat_code = config.FREEBIE_TICKET

    return dict(num_orders=len(orders),
                num_ticket_orders=num_ticket_orders,
                num_new_ticket_orders=num_new_ticket_orders,
                num_existing_ticket_orders=num_existing_ticket_orders,
                num_singing_tickets=num_singing_tickets,
                num_audience_tickets=num_audience_tickets,
                total_tickets=num_audience_tickets+num_singing_tickets,
                event_sku=event_sku,
                event_name=event_name,
                cheat_code=cheat_code
                )




def _process_celebrations(spreadsheet_file, event_date, event_sku):
    SHEET_NAME = 'עסקאות'

    sheet_fields = (
        ORDER_ID,
        FIRST_NAME,
        LAST_NAME,
        PHONE_NUMBER,
        CELEBRATING
    )

    worksheet = load_workbook(spreadsheet_file).get_sheet_by_name(SHEET_NAME)
    column_names = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    column_index_map = {}
    for partial_name in sheet_fields:
        for idx, name in enumerate(column_names):
            if partial_name in name:
                column_index_map[partial_name] = idx

    for row in worksheet.iter_rows(min_row=2, values_only=2):
        order_id = row[column_index_map[ORDER_ID]]
        first_name = row[column_index_map[FIRST_NAME]]
        last_name = row[column_index_map[LAST_NAME]]
        phone_number = row[column_index_map[PHONE_NUMBER]]
        celebrating = row[column_index_map[CELEBRATING]]

        if celebrating:
            Celebration.objects.get_or_create(
                order_id=order_id,
                event_date=event_date,
                event_sku=event_sku,
                customer_name=' '.join([first_name, last_name]),
                phone_number=phone_number,
                celebrating=celebrating
            )
