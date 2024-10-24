"""
Run with a python that has openpyxl installed.
Usage: python analyze_audience [order_data.xlsx] [whatsapp-numbers]

Download a doc from lineup, filtering for as long as you can (a year I think) and for שולם and מומש
Export the following fields (in this order):
לקוח - שם פרטי
לקוח - שם משפחה
שם המוצר
תכונות המוצר - מק״ט
תכונות המוצר - כמות
לקוח - טלפון


Remember to not use events that are currently in the process of being sold!! Incomplete data. Have an exclude list
for this (though sometimes you do want this data)

Need to make adjustments to the script if you want to use several files (since lineup only lets you filter and
download a file for max span of a year)

Looking at people who are strictly audience. If there were a singer at any time, removing them from the accounting
completely

I used the extension "Contacts Extractor for WA" (Paid for a month then canceled) to download all the numbers from both
WhatsApp groups. Saved it in this dir (all-whatsapp-numbers.csv). This script works with the default CSV this extension
generates.
"""
import sys
import re
import csv
from openpyxl import load_workbook
from dataclasses import dataclass, field
from collections import defaultdict
from datetime import datetime

ATTN_SKU = 'ATTN'
SINGER_SKU = 'SING'

def sort_events(event_key):
    date_match = re.search(r'(\d{1,2}\.\d{1,2})', event_key)
    date_str = date_match.group(1)
    date_obj = datetime.strptime(date_str, "%d.%m")
    return date_obj


EXCLUDE_EVENTS = [
    'Broadway With a Twist', # QA night at our house
    # Currently being sold
    # 'Broadway With a Twist - 3.11 - Babu Bar - Tel Aviv',
    # 'Broadway With a Twist - 10.11 - Babu Bar - Tel Aviv',
    # 'Broadway With a Twist - 24.11 - Babu Bar - Tel Aviv',
]

def get_whatsapp_numbers():
    phone_numbers = set()
    phones_path = sys.argv[2]
    with open(phones_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            phone = row.get('Phone')
            if phone:
                phone_numbers.add(phone)

    return phone_numbers

COUNTRY_CODES = ['972', '44', '7']

def get_phone_formats(phone_number):
    phone_number = phone_number.replace(" ", '')
    phone_number = phone_number.lstrip("+")
    phone_number = phone_number.lstrip("0")

    formats = [phone_number]
    for country_code in COUNTRY_CODES:
        formats.append(country_code + phone_number)

    return formats


@dataclass
class AudiencePurchaser:
    events: set = field(default_factory=set)
    total_tickets: int = 0
    was_singer: bool = None
    phone: str = ''

    @property
    def num_times_ordered(self):
        """
        Using a set to remove duplicate events. We're interested in how many different events a person ordered for
        """
        return len(self.events)



audience_purchasers = defaultdict(AudiencePurchaser)
recurring_orders = defaultdict(int)
audience_per_event = defaultdict(int)

if __name__ == "__main__":
    path = sys.argv[1]
    worksheet = load_workbook(path).active
    for row in worksheet.iter_rows(min_row=2, values_only=2):
        first_name, last_name, event_name, ticket_type, total_tickets, phone = row
        name = f"{first_name} {last_name}"
        if ticket_type == ATTN_SKU and name != "אלון אביב" and event_name not in EXCLUDE_EVENTS:
            audience_purchasers[name].total_tickets += total_tickets
            audience_purchasers[name].events.add(event_name)
            audience_purchasers[name].phone = phone
            audience_per_event[event_name] += total_tickets

        if ticket_type == SINGER_SKU and name in audience_purchasers:
            audience_purchasers[name].was_singer = True

    # Delete those who were singers
    audience_purchasers = {name: purchaser for name, purchaser in audience_purchasers.items() if not purchaser.was_singer}

    for name, purchaser in audience_purchasers.items():
        if purchaser.was_singer:
            del audience_purchasers[name]

    for purchaser in audience_purchasers.values():
        recurring_orders[purchaser.num_times_ordered] += 1

    print(f"""
Analyzing {len(audience_per_event)} BWT events.
Total audience orders: {sum(purchaser.num_times_ordered for purchaser in audience_purchasers.values())}
Total audience tickets (each order can heve several tickets): {sum(purchaser.total_tickets for purchaser in audience_purchasers.values())}
Total people that ever ordered an audience ticket: {len(audience_purchasers)}
""")

    for amount_ordered, num_people in sorted(recurring_orders.items()):
        print(f"Amount of people that ordered audience tickets for {amount_ordered} events: {num_people}")

    events_with_single_comers = defaultdict(int)
    for purchaser in audience_purchasers.values():
        if purchaser.num_times_ordered == 1:
            [event] = purchaser.events
            events_with_single_comers[event] += 1

    print("\nHow many people who ordered once (possibly for several people) each event had:")
    for event, num_single_comers in sorted(events_with_single_comers.items(), key=lambda item: sort_events(item[0]),
                                           reverse=True):
        print(f"Event {event}: {num_single_comers}")


    print("\nThe events that repeated comers came to (people who were only audience and never singers):")
    for name, purchaser in audience_purchasers.items():
        if purchaser.num_times_ordered > 1:
            print(f"{name} - Ordered {purchaser.num_times_ordered} times: {purchaser.events}")


    whatsapp_numbers = get_whatsapp_numbers()
    audience_not_in_wa = []
    for purchaser in audience_purchasers.values():
        phone = purchaser.phone
        phone_formats = get_phone_formats(phone)
        for phone_format in phone_formats:
            if phone_format in whatsapp_numbers:
                break
        else:
            audience_not_in_wa.append(phone)

    print(f"\n\n There are {len(audience_not_in_wa)} people who bought audience tickets that "
          f"aren't in the WhatsApp group (out of {len(audience_purchasers)})")
